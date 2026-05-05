from fastapi import FastAPI, HTTPException, Query, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import json
import yfinance as yf
import httpx
import asyncio
import os
from pathlib import Path
from datetime import datetime
from io import BytesIO

try:
    import pandas_ta as ta
except ImportError:
    ta = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

app = FastAPI(title="Equity Analyst API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:1420", "tauri://localhost"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

OLLAMA_BASE = "http://localhost:11434"
OLLAMA_MODEL = "llama3.2:3b"


# ── Models ─────────────────────────────────────────────────────────────────

class StockInfo(BaseModel):
    ticker: str
    name: str
    exchange: str
    sector: str
    market_cap: float
    current_price: float
    previous_close: float
    change: float
    change_pct: float
    pe_ratio: Optional[float]
    eps: Optional[float]
    week_52_high: float
    week_52_low: float
    dividend_yield: Optional[float]
    volume: float
    avg_volume: float
    recommendation: str
    recommendation_mean: Optional[float]
    analyst_buy: int
    analyst_hold: int
    analyst_sell: int
    analyst_total: int
    target_price: Optional[float]
    description: str
    currency: str


class PricePoint(BaseModel):
    date: str
    open: float
    high: float
    low: float
    close: float
    volume: float
    rsi: Optional[float] = None
    macd: Optional[float] = None
    macd_signal: Optional[float] = None
    bb_upper: Optional[float] = None
    bb_middle: Optional[float] = None
    bb_lower: Optional[float] = None


class HistoryResponse(BaseModel):
    ticker: str
    period: str
    data: list[PricePoint]


class SearchResult(BaseModel):
    ticker: str
    name: str
    exchange: str
    type: str


class AnalysisResponse(BaseModel):
    summary: str
    strengths: list[str]
    weaknesses: list[str]
    opportunities: list[str]
    threats: list[str]
    insights: list[str]
    recommendation: str
    valuation: str
    analysis_text: Optional[str] = None


class Message(BaseModel):
    role: str
    content: str


class OllamaStatus(BaseModel):
    ready: bool
    model: str
    message: str


class ChatRequest(BaseModel):
    ticker: str
    question: str
    history: list[Message] = []
    analysis: Optional[dict] = None


class ChatResponse(BaseModel):
    answer: str


# ── Helpers ─────────────────────────────────────────────────────────────────

PERIOD_MAP = {
    "1d":  ("1d",  "5m"),
    "1mo": ("1mo", "1h"),
    "6mo": ("6mo", "1d"),
    "1y":  ("1y",  "1d"),
    "5y":  ("5y",  "1wk"),
}

def safe_float(val) -> Optional[float]:
    try:
        v = float(val)
        return None if (v != v) else v  # NaN check
    except Exception:
        return None

def safe_int(val, default: int = 0) -> int:
    try:
        return int(val)
    except Exception:
        return default


async def call_ollama(messages: list[dict], max_tokens: int = 700) -> str:
    async with httpx.AsyncClient(timeout=60.0) as client:
        res = await client.post(
            f"{OLLAMA_BASE}/api/chat",
            json={
                "model": OLLAMA_MODEL,
                "messages": messages,
                "stream": False,
                "temperature": 0.2,
            },
        )
        res.raise_for_status()
        data = res.json()

    message = data.get("message", {})
    if isinstance(message, dict):
        return message.get("content", "")
    return str(message) if message else ""


def parse_json_from_text(text: str) -> Optional[dict]:
    """Extract and parse JSON from text, handling various formats."""
    import re

    # First try direct JSON parsing
    try:
        return json.loads(text.strip())
    except Exception:
        pass

    # Look for JSON object in text
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end >= 0 and end > start:
        try:
            return json.loads(text[start:end + 1])
        except Exception:
            pass

    # Try to find JSON with regex (more flexible)
    json_pattern = r'\{.*\}'
    match = re.search(json_pattern, text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except Exception:
            pass

    return None


# ── Routes ──────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/search", response_model=list[SearchResult])
async def search_stocks(q: str = Query(..., min_length=1)):
    """Search for stocks by ticker or name using yfinance."""
    try:
        ticker = yf.Ticker(q.upper())
        info = ticker.info

        # yfinance doesn't have a true search endpoint — use basic info lookup
        # In Sprint 2 we can integrate a proper search API; for now return
        # the looked-up ticker if it resolves.
        name = info.get("longName") or info.get("shortName") or q.upper()
        exchange = info.get("exchange", "")
        q_type = info.get("quoteType", "EQUITY")

        results = [SearchResult(
            ticker=q.upper(),
            name=name,
            exchange=exchange,
            type=q_type,
        )]

        # Add common variations to make search feel responsive
        return results

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/stock/{ticker}", response_model=StockInfo)
async def get_stock(ticker: str):
    """Get full stock information."""
    t = ticker.upper()
    try:
        yf_ticker = yf.Ticker(t)
        info = yf_ticker.info

        if not info or "currentPrice" not in info and "regularMarketPrice" not in info:
            raise HTTPException(status_code=404, detail=f"Ticker '{t}' not found.")

        current = safe_float(info.get("currentPrice") or info.get("regularMarketPrice")) or 0.0
        prev_close = safe_float(info.get("previousClose") or info.get("regularMarketPreviousClose")) or current
        change = current - prev_close
        change_pct = (change / prev_close * 100) if prev_close else 0.0

        # Analyst data
        rec_info = info.get("recommendationKey", "hold")
        buy  = safe_int(info.get("numberOfAnalystOpinions", 0))
        # yfinance aggregates — approximate split
        rec_mean = safe_float(info.get("recommendationMean"))
        if rec_mean:
            total = max(buy, 1)
            buy_n  = max(0, round(total * max(0, (3 - rec_mean) / 2)))
            sell_n = max(0, round(total * max(0, (rec_mean - 3) / 2)))
            hold_n = max(0, total - buy_n - sell_n)
        else:
            buy_n = sell_n = hold_n = 0
            total = 0

        return StockInfo(
            ticker=t,
            name=info.get("longName") or info.get("shortName") or t,
            exchange=info.get("exchange") or info.get("fullExchangeName") or "",
            sector=info.get("sector") or info.get("industryDisp") or "Unknown",
            market_cap=safe_float(info.get("marketCap")) or 0.0,
            current_price=current,
            previous_close=prev_close,
            change=change,
            change_pct=change_pct,
            pe_ratio=safe_float(info.get("trailingPE") or info.get("forwardPE")),
            eps=safe_float(info.get("trailingEps")),
            week_52_high=safe_float(info.get("fiftyTwoWeekHigh")) or 0.0,
            week_52_low=safe_float(info.get("fiftyTwoWeekLow")) or 0.0,
            dividend_yield=safe_float(info.get("dividendYield")),
            volume=safe_float(info.get("volume") or info.get("regularMarketVolume")) or 0.0,
            avg_volume=safe_float(info.get("averageVolume")) or 0.0,
            recommendation=rec_info,
            recommendation_mean=rec_mean,
            analyst_buy=buy_n,
            analyst_hold=hold_n,
            analyst_sell=sell_n,
            analyst_total=total,
            target_price=safe_float(info.get("targetMeanPrice")),
            description=info.get("longBusinessSummary") or "",
            currency=info.get("currency") or ("INR" if t.endswith(".NS") or t.endswith(".BO") else "USD"),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/stock/{ticker}/history", response_model=HistoryResponse)
async def get_history(ticker: str, period: str = "1mo"):
    """Get OHLCV price history."""
    t = ticker.upper()
    yf_period, interval = PERIOD_MAP.get(period, ("1mo", "1h"))

    try:
        yf_ticker = yf.Ticker(t)
        hist = yf_ticker.history(period=yf_period, interval=interval)

        if hist.empty:
            raise HTTPException(status_code=404, detail=f"No history for '{t}'.")

        rsi_series = None
        macd_df = None
        bb_df = None
        if ta is not None:
            try:
                rsi_series = ta.rsi(hist["Close"], length=14)
                macd_df = ta.macd(hist["Close"], fast=12, slow=26, signal=9)
                bb_df = ta.bbands(hist["Close"], length=20, std=2)
            except Exception:
                rsi_series = None
                macd_df = None
                bb_df = None

        points: list[PricePoint] = []
        for idx, (ts, row) in enumerate(hist.iterrows()):
            points.append(PricePoint(
                date=str(ts),
                open=round(float(row["Open"]), 4),
                high=round(float(row["High"]), 4),
                low=round(float(row["Low"]), 4),
                close=round(float(row["Close"]), 4),
                volume=float(row["Volume"]),
                rsi=float(rsi_series.iloc[idx]) if rsi_series is not None and rsi_series.iloc[idx] == rsi_series.iloc[idx] else None,
                macd=float(macd_df.iloc[idx]["MACD_12_26_9"]) if macd_df is not None else None,
                macd_signal=float(macd_df.iloc[idx]["MACDs_12_26_9"]) if macd_df is not None else None,
                bb_upper=float(bb_df.iloc[idx]["BBU_20_2.0"]) if bb_df is not None else None,
                bb_middle=float(bb_df.iloc[idx]["BBM_20_2.0"]) if bb_df is not None else None,
                bb_lower=float(bb_df.iloc[idx]["BBL_20_2.0"]) if bb_df is not None else None,
            ))

        return HistoryResponse(ticker=t, period=period, data=points)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ai/analyze", response_model=AnalysisResponse)
async def analyze_stock(ticker: str):
    """Run AI analysis for a stock ticker."""
    t = ticker.upper()
    stock_info = await get_stock(t)
    description = (stock_info.description or "No description available.").strip()
    if len(description) > 1200:
        description = description[:1200] + "..."

    prompt = f"""
You are a senior equity analyst. Produce a concise analysis for the stock {stock_info.ticker} ({stock_info.name}).
Use the facts below and return ONLY valid JSON with these exact keys: summary, strengths, weaknesses, opportunities, threats, insights, recommendation, valuation. RESPOND ONLY WITH VALID JSON. NO OTHER TEXT.

Each SWOT field must be an array of strings. Each insight must be a string. Summary, recommendation, and valuation must be strings.

Example format:
{{
  "summary": "Brief overview of the company and current market position.",
  "strengths": ["Strong brand recognition", "Growing market share"],
  "weaknesses": ["High debt levels", "Dependence on key suppliers"],
  "opportunities": ["Expansion into emerging markets", "New product launches"],
  "threats": ["Increasing competition", "Regulatory changes"],
  "insights": ["Revenue growth has been consistent", "Management has strong track record"],
  "recommendation": "Buy/Hold/Sell",
  "valuation": "Fairly valued/Undervalued/Overvalued"
}}

Facts:
- Exchange: {stock_info.exchange}
- Sector: {stock_info.sector}
- Current price: {stock_info.current_price}
- Previous close: {stock_info.previous_close}
- Change %: {stock_info.change_pct:.2f}%
- Market cap: {stock_info.market_cap}
- P/E: {stock_info.pe_ratio or 'N/A'}
- EPS: {stock_info.eps or 'N/A'}
- 52W high/low: {stock_info.week_52_high}/{stock_info.week_52_low}
- Dividend yield: {stock_info.dividend_yield or 'N/A'}
- Analyst consensus: buy {stock_info.analyst_buy}, hold {stock_info.analyst_hold}, sell {stock_info.analyst_sell}
- Target price: {stock_info.target_price or 'N/A'}
- Business summary: {description}

Write in a confident, actionable tone. Keep the summary and recommendation short. Use bullets for insights.
"""

    messages = [
        {"role": "system", "content": "You are an expert equity analyst specialized in concise investment writeups."},
        {"role": "user", "content": prompt},
    ]

    try:
        answer = await call_ollama(messages)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    parsed = parse_json_from_text(answer)
    if parsed is None:
        # Fallback: try to extract some basic info from the text
        return AnalysisResponse(
            summary=answer.strip()[:500] + "..." if len(answer.strip()) > 500 else answer.strip(),
            strengths=["Analysis generated but SWOT parsing failed"],
            weaknesses=["Analysis generated but SWOT parsing failed"],
            opportunities=["Analysis generated but SWOT parsing failed"],
            threats=["Analysis generated but SWOT parsing failed"],
            insights=["Please check the summary for detailed analysis"],
            recommendation="See summary above",
            valuation="N/A",
            analysis_text=answer.strip(),
        )

    # Ensure all required fields are present and properly typed
    def safe_list(value):
        if isinstance(value, list):
            return [str(item) for item in value if item]
        return []

    def safe_str(value, default=""):
        return str(value) if value else default

    return AnalysisResponse(
        summary=safe_str(parsed.get("summary", ""))[:2000],
        strengths=safe_list(parsed.get("strengths", [])),
        weaknesses=safe_list(parsed.get("weaknesses", [])),
        opportunities=safe_list(parsed.get("opportunities", [])),
        threats=safe_list(parsed.get("threats", [])),
        insights=safe_list(parsed.get("insights", [])),
        recommendation=safe_str(parsed.get("recommendation", "")),
        valuation=safe_str(parsed.get("valuation", "")),
        analysis_text=answer.strip(),
    )


@app.get("/ollama/status", response_model=OllamaStatus)
async def ollama_status():
    """Check if Ollama is running and the model is available."""
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            res = await client.get(f"{OLLAMA_BASE}/api/tags")
            if res.status_code == 200:
                models = [m["name"] for m in res.json().get("models", [])]
                model_ready = any(OLLAMA_MODEL in m for m in models)
                return OllamaStatus(
                    ready=model_ready,
                    model=OLLAMA_MODEL,
                    message="Ready" if model_ready else f"Model {OLLAMA_MODEL} not pulled. Run: ollama pull {OLLAMA_MODEL}",
                )
    except Exception:
        pass

    return OllamaStatus(
        ready=False,
        model=OLLAMA_MODEL,
        message="Ollama is not running. Run: ollama serve",
    )


@app.post("/ai/chat", response_model=ChatResponse)
async def ai_chat(request: ChatRequest):
    if not request.ticker:
        raise HTTPException(status_code=400, detail="Ticker is required.")

    prompt_messages = [
        {
            "role": "system",
            "content": (
                "You are a practical equity research assistant. Use the provided stock analysis and past conversation history to answer follow-up questions clearly "
                "for an investor. If the user asks for a valuation or recommendation, use the analysis context and avoid repeating unrelated details."
            ),
        }
    ]

    if request.analysis is not None:
        analysis_dict = request.analysis if isinstance(request.analysis, dict) else request.analysis.dict()
        analysis_text = (
            f"Analysis summary: {analysis_dict.get('summary', '')}\n"
            f"Strengths: {', '.join(analysis_dict.get('strengths', []))}\n"
            f"Weaknesses: {', '.join(analysis_dict.get('weaknesses', []))}\n"
            f"Opportunities: {', '.join(analysis_dict.get('opportunities', []))}\n"
            f"Threats: {', '.join(analysis_dict.get('threats', []))}\n"
            f"Insights: {', '.join(analysis_dict.get('insights', []))}\n"
            f"Recommendation: {analysis_dict.get('recommendation', '')}\n"
            f"Valuation: {analysis_dict.get('valuation', '')}\n"
        )
        prompt_messages.append({"role": "system", "content": analysis_text})

    for message in request.history:
        if message.role in {"user", "assistant"}:
            prompt_messages.append({"role": message.role, "content": message.content})

    prompt_messages.append({"role": "user", "content": request.question})

    try:
        answer = await call_ollama(prompt_messages)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


    return ChatResponse(answer=answer.strip())


# ── Portfolio Routes ─────────────────────────────────────────────────────────

class PortfolioPosition(BaseModel):
    id: str
    ticker: str
    quantity: float
    buy_price: float
    current_price: Optional[float] = None
    sector: Optional[str] = None
    name: Optional[str] = None


class PortfolioData(BaseModel):
    positions: list[PortfolioPosition]
    total_invested: float
    total_current_value: float
    total_gain_loss: float
    total_gain_loss_pct: float
    sector_allocation: dict[str, float]


class PortfolioAnalysisResponse(BaseModel):
    summary: str
    risks: list[str]
    opportunities: list[str]
    sector_balance: str
    concentration: str
    recommendations: list[str]
    suggested_rebalance: Optional[str] = None


class PortfolioPriceRequest(BaseModel):
    tickers: list[str]


@app.post("/portfolio/prices")
async def get_portfolio_prices(request: PortfolioPriceRequest):
    """Fetch current prices for portfolio tickers."""
    result = {}

    for ticker in request.tickers:
        try:
            yf_ticker = yf.Ticker(ticker)
            info = yf_ticker.info
            current = safe_float(info.get("currentPrice") or info.get("regularMarketPrice")) or 0.0
            sector = info.get("sector") or "Unknown"
            name = info.get("longName") or ticker

            result[ticker] = {
                "price": current,
                "sector": sector,
                "name": name,
            }
        except Exception:
            result[ticker] = {
                "price": 0.0,
                "sector": "Unknown",
                "name": ticker,
            }

    return result


@app.post("/portfolio/analyze", response_model=PortfolioAnalysisResponse)
async def analyze_portfolio(portfolio: PortfolioData):
    """Analyze portfolio for concentration risk, sector balance, and rebalancing."""
    total_value = portfolio.total_current_value or 1.0

    # Calculate concentration
    top_3_value = sorted(
        [p.quantity * (p.current_price or p.buy_price) for p in portfolio.positions],
        reverse=True,
    )[:3]
    top_3_pct = sum(top_3_value) / total_value * 100 if total_value > 0 else 0

    # Sector concentration
    sector_details = []
    for sector, value in sorted(
        portfolio.sector_allocation.items(), key=lambda x: x[1], reverse=True
    )[:3]:
        pct = (value / total_value * 100) if total_value > 0 else 0
        sector_details.append(f"{sector}: {pct:.1f}%")

    prompt = f"""
You are a portfolio advisor. Analyze this portfolio and return ONLY valid JSON with keys: summary, risks, opportunities, sector_balance, concentration, recommendations, suggested_rebalance.

Portfolio Summary:
- Total Invested: {portfolio.total_invested}
- Current Value: {portfolio.total_current_value}
- Gain/Loss: {portfolio.total_gain_loss} ({portfolio.total_gain_loss_pct:.2f}%)
- Top 3 holdings: {top_3_pct:.1f}% of portfolio
- Sector breakdown: {', '.join(sector_details)}
- Number of holdings: {len(portfolio.positions)}

Example format:
{{
  "summary": "Brief portfolio overview",
  "risks": ["Risk 1", "Risk 2"],
  "opportunities": ["Opportunity 1"],
  "sector_balance": "Assessment of sector diversification",
  "concentration": "Assessment of concentration risk",
  "recommendations": ["Recommendation 1"],
  "suggested_rebalance": "Specific rebalancing suggestion or null"
}}

Provide concise, actionable advice for an Indian investor.
"""

    messages = [
        {
            "role": "system",
            "content": "You are an expert portfolio advisor specializing in Indian equity markets.",
        },
        {"role": "user", "content": prompt},
    ]

    try:
        answer = await call_ollama(messages)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    parsed = parse_json_from_text(answer)
    if parsed is None:
        return PortfolioAnalysisResponse(
            summary=answer.strip()[:500],
            risks=["Analysis generated but parsing failed"],
            opportunities=[],
            sector_balance="See summary",
            concentration="See summary",
            recommendations=[],
        )

    def safe_list(value):
        if isinstance(value, list):
            return [str(item) for item in value if item]
        return []

    def safe_str(value, default=""):
        return str(value) if value else default

    return PortfolioAnalysisResponse(
        summary=safe_str(parsed.get("summary", ""))[:2000],
        risks=safe_list(parsed.get("risks", [])),
        opportunities=safe_list(parsed.get("opportunities", [])),
        sector_balance=safe_str(parsed.get("sector_balance", "")),
        concentration=safe_str(parsed.get("concentration", "")),
        recommendations=safe_list(parsed.get("recommendations", [])),
        suggested_rebalance=safe_str(parsed.get("suggested_rebalance")),
    )


class PortfolioChatRequest(BaseModel):
    question: str
    history: list[Message] = []
    portfolio: PortfolioData
    analysis: Optional[dict] = None


# ── NetWorth Models ────────────────────────────────────────────────────────

class NetWorthInvestment(BaseModel):
    id: str
    ticker: str
    quantity: float
    current_price: Optional[float] = None
    purchase_date: Optional[str] = None
    notes: Optional[str] = None


class NetWorthOtherAsset(BaseModel):
    id: str
    name: str
    amount: float
    type: str  # RD, FD, Gold, Crypto, etc.
    purchase_date: Optional[str] = None
    notes: Optional[str] = None


class NetWorthData(BaseModel):
    bank_balance: float
    investments: list[NetWorthInvestment] = []
    other_assets: list[NetWorthOtherAsset] = []
    last_updated: str = ""


NETWORTH_DATA_FILE = Path.home() / ".equity-analyst" / "networth.json"


def ensure_networth_dir():
    """Ensure the networth data directory exists."""
    NETWORTH_DATA_FILE.parent.mkdir(parents=True, exist_ok=True)


def load_networth_data() -> NetWorthData:
    """Load networth data from file."""
    ensure_networth_dir()
    if NETWORTH_DATA_FILE.exists():
        try:
            with open(NETWORTH_DATA_FILE, "r") as f:
                data = json.load(f)
                return NetWorthData(**data)
        except Exception as e:
            print(f"Error loading networth data: {e}")
    return NetWorthData(bank_balance=0.0)


def save_networth_data(data: NetWorthData):
    """Save networth data to file."""
    ensure_networth_dir()
    data.last_updated = datetime.now().isoformat()
    with open(NETWORTH_DATA_FILE, "w") as f:
        json.dump(data.dict(), f, indent=2)


@app.post("/portfolio/chat")
async def portfolio_chat(request: PortfolioChatRequest):
    """Chat about portfolio analysis."""
    messages = [
        {
            "role": "system",
            "content": "You are a portfolio advisor helping with questions about portfolio composition, diversification, and rebalancing.",
        }
    ]

    # Add portfolio context
    if request.analysis:
        analysis_text = (
            f"Portfolio Analysis:\n"
            f"Summary: {request.analysis.get('summary', '')}\n"
            f"Risks: {', '.join(request.analysis.get('risks', []))}\n"
            f"Recommendations: {', '.join(request.analysis.get('recommendations', []))}\n"
        )
        messages.append({"role": "system", "content": analysis_text})

    # Add chat history
    for msg in request.history:
        if msg.role in {"user", "assistant"}:
            messages.append({"role": msg.role, "content": msg.content})

    messages.append({"role": "user", "content": request.question})

    try:
        answer = await call_ollama(messages)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return ChatResponse(answer=answer.strip())


# ── NetWorth Routes ────────────────────────────────────────────────────────

@app.get("/networth/data")
async def get_networth_data():
    """Get current networth data."""
    data = load_networth_data()
    
    # Fetch current prices for investments
    if data.investments:
        tickers = [inv.ticker for inv in data.investments]
        try:
            for inv in data.investments:
                yf_ticker = yf.Ticker(inv.ticker)
                inv.current_price = safe_float(
                    yf_ticker.info.get("currentPrice") or 
                    yf_ticker.info.get("regularMarketPrice")
                ) or inv.current_price or 0.0
        except Exception as e:
            print(f"Error fetching prices: {e}")
    
    return data


@app.post("/networth/save")
async def save_networth(data: NetWorthData):
    """Save networth data."""
    try:
        save_networth_data(data)
        return {"status": "ok", "message": "Networth data saved"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/networth/valuations")
async def get_networth_valuations(tickers: list[str]):
    """Get current valuations for multiple tickers."""
    valuations = {}
    for ticker in tickers:
        try:
            yf_ticker = yf.Ticker(ticker.upper())
            price = safe_float(
                yf_ticker.info.get("currentPrice") or 
                yf_ticker.info.get("regularMarketPrice")
            ) or 0.0
            valuations[ticker.upper()] = {
                "price": price,
                "currency": yf_ticker.info.get("currency", "INR"),
                "name": yf_ticker.info.get("longName") or ticker.upper()
            }
        except Exception as e:
            valuations[ticker.upper()] = {"error": str(e), "price": None}
    
    return valuations


@app.post("/networth/export")
async def export_networth_excel():
    """Export networth data to Excel file."""
    if not Workbook:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    data = load_networth_data()
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Net Worth"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Summary section
        ws['A1'] = "Net Worth Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Bank Balance
        ws['A4'] = "Bank Balance"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = data.bank_balance
        ws['B4'].number_format = '#,##0.00'
        
        # Investments section
        row = 6
        ws[f'A{row}'] = "Investments"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        headers = ['Ticker', 'Quantity', 'Current Price', 'Value', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        total_investment_value = 0
        for inv in data.investments:
            current_price = inv.current_price or 0.0
            value = inv.quantity * current_price
            total_investment_value += value
            
            ws.cell(row=row, column=1).value = inv.ticker
            ws.cell(row=row, column=2).value = inv.quantity
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3).value = current_price
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4).value = value
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5).value = inv.notes or ""
            row += 1
        
        # Total row
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4).value = total_investment_value
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        
        # Other Assets section
        row += 3
        ws[f'A{row}'] = "Other Assets (RD, FD, etc.)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        headers = ['Type', 'Amount', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        total_other_assets = 0
        for asset in data.other_assets:
            total_other_assets += asset.amount
            ws.cell(row=row, column=1).value = f"{asset.type}: {asset.name}"
            ws.cell(row=row, column=2).value = asset.amount
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3).value = asset.notes or ""
            row += 1
        
        # Total row
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = total_other_assets
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        
        # Summary at bottom
        row += 3
        ws[f'A{row}'] = "Total Net Worth Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws[f'A{row}'] = "Bank Balance:"
        ws[f'B{row}'] = data.bank_balance
        ws[f'B{row}'].number_format = '#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Investments Value:"
        ws[f'B{row}'] = total_investment_value
        ws[f'B{row}'].number_format = '#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Other Assets:"
        ws[f'B{row}'] = total_other_assets
        ws[f'B{row}'].number_format = '#,##0.00'
        
        row += 1
        total_networth = data.bank_balance + total_investment_value + total_other_assets
        ws[f'A{row}'] = "TOTAL NET WORTH"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws[f'B{row}'] = total_networth
        ws[f'B{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws[f'B{row}'].number_format = '#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        # Save to file
        file_path = Path.home() / ".equity-analyst" / f"networth_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ensure_networth_dir()
        wb.save(file_path)
        
        return {
            "status": "ok",
            "message": "Excel file created",
            "file_path": str(file_path),
            "total_networth": total_networth
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
